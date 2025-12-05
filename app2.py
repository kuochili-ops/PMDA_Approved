
import streamlit as st
import pandas as pd
import requests
import re
import time
import os
import openpyxl

# ====== API 金鑰設定 ======
AZURE_KEY = st.secrets["AZURE_KEY"]
AZURE_REGION = st.secrets["AZURE_REGION"]
endpoint = "https://api.cognitive.microsofttranslator.com/translate"
headers = {
    "Ocp-Apim-Subscription-Key": AZURE_KEY,
    "Ocp-Apim-Subscription-Region": AZURE_REGION,
    "Content-type": "application/json"
}

# ====== KEGG API 查詢函式 ======
def kegg_drug_english_names(jp_name):
    url = f"https://rest.kegg.jp/find/drug/{jp_name}"
    try:
        resp = requests.get(url, timeout=10)
        if resp.ok and resp.text:
            line = resp.text.split('\n')[0]
            fields = line.split()
            if len(fields) > 1:
                names = [n.strip() for n in fields[1].split(';')]
                trade_names = [n for n in names if n and n[0].isupper() and not n.isupper()]
                english_names = [n for n in names if n and n[0].isupper() and n.isalpha()]
                return {
                    "trade_name_en_kegg": trade_names[0] if trade_names else "",
                    "ingredient_en_kegg": english_names[0] if english_names else ""
                }
    except Exception:
        pass
    return {"trade_name_en_kegg": "", "ingredient_en_kegg": ""}

# ====== Microsoft Translator API 單句翻譯 ======
def ms_translator(text, from_lang="ja"):
    body = [{"text": text}]
    params = {"api-version": "3.0", "from": from_lang, "to": ["en"]}
    try:
        resp = requests.post(endpoint, params=params, headers=headers, json=body, timeout=10)
        if resp.ok:
            data = resp.json()
            return data[0]["translations"][0]["text"]
    except Exception:
        pass
    return ""

# ====== 取得分頁列印範圍 ======
def get_print_area(sheet):
    area = sheet.print_area
    if area:
        # openpyxl 3.1+ print_area 會是 tuple
        if isinstance(area, (list, tuple)):
            area = area[0]
        return area
    return None

def read_print_area_to_df(xls_path, sheet_name):
    wb = openpyxl.load_workbook(xls_path, data_only=True)
    ws = wb[sheet_name]
    area = get_print_area(ws)
    if not area:
        return None  # 沒有設定列印範圍
    # 處理多個範圍，只取第一個
    area = str(area).split(',')[0]
    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(area)
    except ValueError:
        return None
    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        data.append(row)
    if not data or len(data) < 2:
        return None
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

# ====== 資料清理函式（強化版） ======
def clean_dataframe(df):
    if not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    rename_map = {}
    for col in df.columns:
        if re.match(r'^販.*売.*名.*', str(col)):
            rename_map[col] = '販賣名/公司 (日文)'
        elif re.match(r'^成.*分.*名.*', str(col)):
            rename_map[col] = '成分名 (日文)'
        elif re.match(r'^No\\.?$', str(col)):
            rename_map[col] = 'No.'
    df = df.rename(columns=rename_map)
    # 只保留有藥品編號、販賣名、成分名的行
    if {'No.', '販賣名/公司 (日文)', '成分名 (日文)'}.issubset(df.columns):
        df = df[
            df['No.'].apply(lambda x: str(x).strip().isdigit()) &
            df['販賣名/公司 (日文)'].astype(str).str.strip().ne('') &
            df['成分名 (日文)'].astype(str).str.strip().ne('')
        ]
    elif '成分名 (日文)' in df.columns:
        df = df[df['成分名 (日文)'].notnull() & (df['成分名 (日文)'].astype(str).str.strip() != '')]
    else:
        df = pd.DataFrame()  # 沒有主要欄位就回傳空表
    # 去除全空白行
    if not df.empty:
        df = df.dropna(how='all')
        df = df[~(df.applymap(lambda x: str(x).strip() == '').all(axis=1))]
        df = df.reset_index(drop=True)
    return df

# ====== 分頁另存 CSV（只處理列印範圍） ======
def save_sheets_to_csv_by_print_area(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_map = {}
    for sheet_name in wb.sheetnames:
        df = read_print_area_to_df(uploaded_file, sheet_name)
        if df is None or df.empty:
            st.write(f"分頁「{sheet_name}」無有效資料或未設定列印範圍/格式異常，已跳過。")
            continue
        raw_count = len(df)
        df = clean_dataframe(df)
        clean_count = len(df)
        # 嘗試找月份
        month_match = re.search(r'(\d+)月', sheet_name)
        if not month_match:
            for col in df.columns:
                m = re.search(r'(\d+)月', str(col))
                if m:
                    month_match = m
                    break
        if month_match:
            month = month_match.group(1) + "月"
        else:
            month = sheet_name
        csv_name = f"{month}.csv"
        df.to_csv(csv_name, index=False, encoding="utf-8")
        sheet_map[month] = (csv_name, raw_count, clean_count)
    return sheet_map

# ====== 翻譯主流程 ======
def translate_and_combine(df):
    st.write(f"清理後有效資料共 {len(df)} 筆")
    trade_name_en_list = []
    ingredient_en_list = []
    progress = st.empty()
    for idx, row in df.iterrows():
        progress.info(f"第 {idx+1} 項翻譯中…")
        kegg_result = kegg_drug_english_names(str(row.get('販賣名/公司 (日文)', '')))
        trade_name_en = kegg_result["trade_name_en_kegg"]
        ingredient_en = kegg_result["ingredient_en_kegg"]
        if not trade_name_en:
            trade_name_en = ms_translator(str(row.get('販賣名/公司 (日文)', '')))
        if not ingredient_en:
            ingredient_en = ms_translator(str(row.get('成分名 (日文)', '')))
        trade_name_en_list.append(trade_name_en)
        ingredient_en_list.append(ingredient_en)
        time.sleep(0.34)  # KEGG 頻率限制
    progress.success("全部翻譯完成！")
    df['Trade Name/Company (English)'] = trade_name_en_list
    df['Ingredient Name (English)'] = ingredient_en_list
    return df

# ====== Streamlit 主程式 ======
def main():
    st.set_page_config(layout="wide", page_title="PMDA 日本新藥翻譯列表生成器")
    st.title("🇯🇵 PMDA 日本新藥翻譯列表生成器 (自動分頁轉 CSV + 翻譯)")
    uploaded_file = st.file_uploader("上傳 PMDA 公告 Excel 檔案", type=['xlsx', 'xls'])
    if uploaded_file:
        st.info("正在自動分割各月份（僅處理分頁列印範圍）...")
        month_csv_map = save_sheets_to_csv_by_print_area(uploaded_file)
        if not month_csv_map:
            st.warning("未偵測到任何有效分頁或分頁未設定列印範圍。")
            return
        for month, (csv_name, raw_count, clean_count) in month_csv_map.items():
            st.subheader(f"{month} 翻譯結果")
            st.write(f"列印範圍原始筆數：{raw_count}，清理後：{clean_count}")
            df = pd.read_csv(csv_name, encoding="utf-8")
            if df.empty:
                st.warning(f"{month} 無有效資料，已跳過。")
                continue
            translated_df = translate_and_combine(df)
            st.dataframe(translated_df, use_container_width=True, hide_index=True)
            csv_export = translated_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label=f"📥 下載 {month} 翻譯結果 (CSV)",
                data=csv_export,
                file_name=f"{month}_Translated.csv",
                mime='text/csv'
            )
            os.remove(csv_name)

if __name__ == "__main__":
    main()
